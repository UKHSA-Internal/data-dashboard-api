import datetime
from pathlib import Path

from django.db import models
from openpyxl import load_workbook

FILE_PATH = f"{Path(__file__).resolve().parent.parent}/migration_data/"
FILE_NAME = "metrics_definitions_migration.xlsx"

metric_docs_workbook = load_workbook(filename=f"{FILE_PATH}{FILE_NAME}", read_only=True)
work_sheet = metric_docs_workbook.active


def build_sections(sections) -> list[dict]:
    """Build metric documentation page sections.

    Args:
        sections: list of tuples containing section details.

    Returns:
        list of dictionaries containing metric documentation
        section details.
    """
    return [
        {
            "type": "section",
            "value": {
                "title": section[0],
                "body": section[1],
            },
            "id": models.BigAutoField(
                    auto_created=True,
                    primary_key=True,
                    serialize=False,
                    verbose_name="ID",
                ),
        } for section in sections
    ]


def build_entry_from_row_data(row) -> dict[str | list[dict]]:
    """Build a metrics documentation page entry.

    Args:
        row: tuple containing spreadsheet row data.

    Returns:
        dictionary containing metric documentation entry.
    """
    return {
        "title": row[5],
        "depth": 1,
        "path": "abc",
        "date_posted": datetime.datetime.today(),
        "page_description": row[2],
        "metric": row[2],
        "metric_group": row[1].lower(),
        "topic": row[0],
        "body": build_sections(
            [
                ("rationale", row[3]),
                ("definition", row[4]),
                ("methodology", row[6]),
                ("caveats", row[7]),
            ]
        ),
    }


def get_metrics_definitions() -> list[dict]:
    """Retrieves a list of metrics documentation entries from a spreadsheet.

    Returns:
        list of dictionaries containing metric documentation page entries.
    """
    return [
        build_entry_from_row_data(row)
        for row in work_sheet.iter_rows(min_row=2, max_row=3, values_only=True)
    ]

data = get_metrics_definitions()
print(data)
