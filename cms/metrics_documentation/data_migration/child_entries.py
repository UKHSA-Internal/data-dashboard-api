import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def build_sections(*, sections: list[tuple[str, str]]) -> list[dict]:
    """Build metric documentation page sections.

    Args:
        sections: list of tuples containing section details.

    Returns:
        list of dictionaries containing metric documentation
        section details.
    """
    return [
        {
            "type": "section",
            "value": {
                "title": section[0],
                "body": section[1],
            },
        }
        for section in sections
    ]


def build_entry_from_row_data(*, row: tuple[str, ...]) -> dict[str, str | list[dict]]:
    """Build a metrics documentation page entry.

    Args:
        row: tuple containing spreadsheet row data.

    Returns:
        dictionary containing metric documentation entry.
    """
    title: str = row[0]
    page_description: str = row[4]
    metric: str = row[1]
    sections: list[tuple[str, str]] = gather_sections_and_omit_if_needed(row=row)
    return {
        "title": title,
        "seo_title": f"{title} | UKHSA data dashboard",
        "search_description": page_description,
        "page_description": page_description,
        "metric": metric,
        "body": build_sections(sections=sections),
    }


def gather_sections_and_omit_if_needed(
    *, row: tuple[str, ...]
) -> list[tuple[str, str]]:
    """Builds a list of sections which can be parsed by the `build_sections()` function

    Notes:
        If an accompanying body of text for the section is an empty string,
        then the section will be omitted entirely.

    Args:
        row: Tuple of strings representing each column value
            of the given child entry

    Returns:
        List of tuples whereby each section consists of
            1) The section title
            2) The body of text for the associated section

    """
    title_indexes = [
        ("Rationale", 2),
        ("Definition", 3),
        ("Methodology", 5),
        ("Caveats", 6),
    ]

    return [(title, body) for title, index in title_indexes if (body := row[index])]


def _load_source_data_as_worksheet() -> Worksheet:
    file_path = f"{Path(__file__).resolve().parent.parent}/data_migration/source_data/"
    file_name = "metrics_definitions_migration_edit.xlsx"

    metric_docs_workbook = load_workbook(
        filename=f"{file_path}{file_name}", read_only=True
    )
    return metric_docs_workbook.active


def get_metrics_definitions() -> list[dict[str | list[dict]]]:
    """Retrieves a list of metrics documentation entries from a spreadsheet.

    Returns:
        list of dictionaries containing metric documentation page entries.
    """
    work_sheet = _load_source_data_as_worksheet()
    return [
        build_entry_from_row_data(row=row)
        for row in work_sheet.iter_rows(min_row=2, values_only=True)
    ]
