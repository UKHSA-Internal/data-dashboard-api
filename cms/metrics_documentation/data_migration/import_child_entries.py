import datetime
from pathlib import Path

from openpyxl import load_workbook


def build_sections(sections) -> list[dict]:
    """Build metric documentation page sections.

    Args:
        sections: list of tuples containing section details.

    Returns:
        list of dictionaries containing metric documentation
        section details.
    """
    return [
        {
            "type": "section",
            "value": {
                "title": section[0],
                "body": section[1],
            },
        }
        for section in sections
    ]


def build_entry_from_row_data(row) -> dict[str | list[dict]]:
    """Build a metrics documentation page entry.

    Args:
        row: tuple containing spreadsheet row data.

    Returns:
        dictionary containing metric documentation entry.
    """
    return {
        "title": row[0],
        "date_posted": datetime.datetime.today(),
        "page_description": row[4],
        "metric": row[1],
        "body": build_sections(
            [
                ("Rationale", row[2]),
                ("Definition", row[3]),
                ("Methodology", row[5]),
                ("Caveats", row[6]),
            ]
        ),
    }


def _load_worksheet():
    file_path = f"{Path(__file__).resolve().parent.parent}/migration_data/"
    file_name = "metrics_definitions_migration_edit.xlsx"

    metric_docs_workbook = load_workbook(
        filename=f"{file_path}{file_name}", read_only=True
    )
    return metric_docs_workbook.active


def get_metrics_definitions() -> list[dict]:
    """Retrieves a list of metrics documentation entries from a spreadsheet.

    Returns:
        list of dictionaries containing metric documentation page entries.
    """
    work_sheet = _load_worksheet()
    return [
        build_entry_from_row_data(row)
        for row in work_sheet.iter_rows(min_row=2, max_row=28, values_only=True)
    ]
