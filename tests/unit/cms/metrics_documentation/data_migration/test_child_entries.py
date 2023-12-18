import datetime
from unittest import mock

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cms.metrics_documentation.data_migration.child_entries import (
    _load_source_data_as_worksheet,
    build_entry_from_row_data,
    build_sections,
    get_metrics_definitions,
)
from metrics.api.settings import ROOT_LEVEL_BASE_DIR

MODULE_PATH = "cms.metrics_documentation.data_migration.child_entries"


class TestBuildSections:
    def test_returns_correct_list(self):
        """
        Given a list of tuples representing
            the title and body of each section
        When `build_sections()` is called
        Then the correct output is returned
        """
        # Given
        sections = [
            ("Rationale", "Fake rationale content"),
            ("Definition", "Fake definition content"),
            ("Methodology", "Fake methodology content"),
            ("Caveats", "Fake caveats content"),
        ]

        # When
        constructed_sections = build_sections(sections=sections)

        # Then
        for section in sections:
            expected_section = {
                "type": "section",
                "value": {"title": section[0], "body": section[1]},
            }
            assert expected_section in constructed_sections


class TestBuildEntryFromRowData:
    @mock.patch(f"{MODULE_PATH}.build_sections")
    def test_returns_correct_dictionary(
        self,
        spy_build_sections: mock.MagicMock,
    ):
        """
        Given a tuple representing a metrics documentation
            child entry.
        When `build_entry_from_row_data()` is called.
        Then the correct output is returned.
        """
        # Given
        fake_row = (
            "Fake title",
            "Fake_metric_name",
            "Fake rationale content",
            "Fake definition content",
            "Fake page description",
            "Fake methodology content",
            "Fake caveats content",
        )
        spy_build_sections.return_value = []
        expected_response = {
            "title": "Fake title",
            "seo_title": "Fake title | UKHSA data dashboard",
            "search_description": "Fake page description",
            "date_posted": datetime.datetime.today().strftime("%Y-%m-%d"),
            "page_description": "Fake page description",
            "metric": "Fake_metric_name",
            "body": [],
        }

        # When
        response = build_entry_from_row_data(fake_row)

        # Then
        assert response == expected_response


class TestGetMetricsDefinitions:
    @staticmethod
    def build_worksheet() -> Worksheet:
        """creates and returns an openpyxl worksheet.

        Returns:
            Worksheet
        """
        work_book = Workbook()
        work_sheet = work_book.active

        work_sheet["A2"] = "Fake title"
        work_sheet["B2"] = "Fake_metric_name"
        work_sheet["C2"] = "Fake rationale content"
        work_sheet["D2"] = "Fake definition content"
        work_sheet["E2"] = "Fake page description"
        work_sheet["F2"] = "Fake methodology content"
        work_sheet["G2"] = "Fake caveats content"

        return work_sheet

    @mock.patch(f"{MODULE_PATH}._load_source_data_as_worksheet")
    def test_delegates_calls_correctly(
        self,
        spy_load_source_data_as_worksheet: mock.MagicMock,
    ):
        """
        Given a fake Excel spreadsheet (openpyxl worksheet).
        When the `get_metrics_definition()` method is called.
        Then the expected response should be returned.
        """
        # Given
        spy_load_source_data_as_worksheet.return_value = self.build_worksheet()
        expected_response = [
            {
                "title": "Fake title",
                "seo_title": "Fake title | UKHSA data dashboard",
                "search_description": "Fake page description",
                "date_posted": datetime.datetime.today().strftime("%Y-%m-%d"),
                "page_description": "Fake page description",
                "metric": "Fake_metric_name",
                "body": [
                    {
                        "type": "section",
                        "value": {
                            "body": "Fake rationale content",
                            "title": "Rationale",
                        },
                    },
                    {
                        "type": "section",
                        "value": {
                            "body": "Fake definition content",
                            "title": "Definition",
                        },
                    },
                    {
                        "type": "section",
                        "value": {
                            "body": "Fake methodology content",
                            "title": "Methodology",
                        },
                    },
                    {
                        "type": "section",
                        "value": {"body": "Fake caveats content", "title": "Caveats"},
                    },
                ],
            }
        ]

        # When
        response = get_metrics_definitions()

        # Then
        assert response == expected_response


class TestLoadSourceDataAsWorksheet:
    @mock.patch(f"{MODULE_PATH}.load_workbook")
    def test_delegates_call_with_correct_file_path(
        self, spy_load_workbook: mock.MagicMock
    ):
        """
        Given no input
        When `_load_source_data_as_worksheet()` is called
        Then the call is delegated to the `load_workbook()`
            function with the correct path to the source data sheet

        Patches:
            `spy_load_workbook`: For the main assertion
        """
        # Given / When
        worksheet = _load_source_data_as_worksheet()

        # Then
        assert worksheet == spy_load_workbook.return_value.active
        expected_filename: str = (
            f"{ROOT_LEVEL_BASE_DIR}/cms/metrics_documentation/data_migration/source_data"
            f"/metrics_definitions_migration_edit.xlsx"
        )
        spy_load_workbook.assert_called_with(filename=expected_filename, read_only=True)
